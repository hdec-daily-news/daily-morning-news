# -*- coding: utf-8 -*-
"""
data/links.json → 엑셀 파일 생성 (프로젝트1 산출물)
기존 참고 양식(260708__일일 정치 정당 경제 주요 기사 정리_송부.xlsx)의 보안 해제 사본을 받으면
이 스크립트의 서식 부분을 그 양식에 맞춰 조정한다.
"""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

SECTOR_TITLES = {
    "politics_main": "1) 정치 관련 주요기사",
    "dp": "2) 더불어민주당 동정",
    "ppp": "3) 국민의힘 동정",
    "economy": "4) 경제 동정",
}


def build(data, out_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "일일정치주요뉴스"

    ws["A1"] = f"일일 정치 주요뉴스 {data['date_label']}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws["A2"] = f"(기준: {data['window_start'][:16].replace('T', ' ')} ~ {data['window_end'][:16].replace('T', ' ')})"
    ws["A2"].font = Font(name="Arial", size=9, italic=True)

    row = 4
    for sector in data["sectors"]:
        title = SECTOR_TITLES.get(sector["key"], sector["label"])
        ws.cell(row=row, column=1, value=title).font = Font(name="Arial", bold=True, size=12)
        row += 1
        if not sector["articles"]:
            ws.cell(row=row, column=1, value="(해당 시간대 기사 없음)").font = Font(
                name="Arial", italic=True, color="999999"
            )
            row += 1
        for a in sector["articles"]:
            c1 = ws.cell(row=row, column=1, value=a["title"])
            c1.font = Font(name="Arial", size=11)
            c1.alignment = Alignment(wrap_text=True, vertical="top")
            c2 = ws.cell(row=row, column=2, value=a["link"])
            c2.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 65
    ws.column_dimensions["B"].width = 55

    wb.save(out_path)


def main():
    with open("data/links.json", encoding="utf-8") as f:
        data = json.load(f)
    os.makedirs("output", exist_ok=True)
    date_compact = data["window_end"][:10].replace("-", "")[2:]  # e.g. 260709
    out_path = f"output/{date_compact}__일일_정치_정당_경제_주요_기사_정리.xlsx"
    build(data, out_path)
    # 최신본 고정 파일명도 함께 생성 (페이지 다운로드 링크용)
    build(data, "output/latest.xlsx")
    print(f"[OK] 엑셀 생성: {out_path}")


if __name__ == "__main__":
    main()
